#!/usr/bin/env python3
"""
Extract weekly turf-maintenance percentages from the input workbook into JSON.

The maintenance team's workbook (one file, named turf-maintenance_YYYY[-MM].xlsx)
has one tab PER MONTH (April..December); the previous per-district "Location"
tabs were removed. Each month tab lists every zone/reach across all three
districts as rows, with weekly Cycle 1 / Cycle 2 % columns G/J/M/P/S and
H/K/N/Q/T. We read the latest month tab that has data and take the MAX
cumulative % per reach across its weekly columns.

Output: src/data/turfCycles.json -- a flat list of {zone, reach, cycle1Pct,
cycle2Pct, c1CompletedOn, c2CompletedOn} plus the reporting month and any
warnings. grassCutting.ts matches these onto the curated zones by normalized
zone+reach name, and only when the reporting month is newer than its curated
baseline (so incomplete data never regresses).

Which month gets credit is decided by which TAB the work is entered on, not by
the completion dates -- those are a cross-check. Each month tab's final column
ends on the last calendar day of the month (Aug 2026 onward) so month-end work
lands on the right tab in the first place; a completion date that disagrees with
its tab is reported as a warning.

Usage:
  python3 scripts/extractTurfData.py [path-to-workbook.xlsx]
  TURF_INPUT=<path> python3 scripts/extractTurfData.py
"""
import datetime
import json
import os
import re
import sys

import openpyxl

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE_DIR = os.path.join(REPO_ROOT, "data/sources/turf")
OUTPUT_PATH = os.path.join(REPO_ROOT, "src/data/turfCycles.json")
DEFAULT_FILENAME = "turf-maintenance.xlsx"

MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
}

# 0-indexed: Zone=A(0), Reach=E(4); weekly C1 = G,J,M,P,S; C2 = H,K,N,Q,T.
# These positions are FIXED. Inserting a column anywhere at or before T shifts
# every reading silently -- no error, just wrong numbers -- so the template's
# instructions tab tells the team to append rather than insert.
ZONE_COL, REACH_COL = 0, 4
C1_COLS = [6, 9, 12, 15, 18]
C2_COLS = [7, 10, 13, 16, 19]
# Per-cycle completion dates: W(22) = Cycle 1, X(23) = Cycle 2. Added Aug 2026
# after month-end work was being credited to the wrong month (V is "Updated By").
C1_DATE_COL, C2_DATE_COL = 22, 23


def resolve_input():
    if len(sys.argv) > 1:
        return os.path.expanduser(sys.argv[1])
    if os.environ.get("TURF_INPUT"):
        return os.path.expanduser(os.environ["TURF_INPUT"])
    return os.path.join(BASE_DIR, DEFAULT_FILENAME)


def max_pct(row, cols):
    vals = [float(row[c]) for c in cols if c < len(row) and isinstance(row[c], (int, float))]
    return max(vals) if vals else None


def cell_date(row, col):
    """Return (ISO date string or None, parsed_ok) for a completion-date cell."""
    if col >= len(row) or row[col] is None or str(row[col]).strip() == "":
        return None, True
    v = row[col]
    if isinstance(v, datetime.datetime):
        return v.date().isoformat(), True
    if isinstance(v, datetime.date):
        return v.isoformat(), True
    return str(v).strip(), False  # typed as text, e.g. "6/30" -- can't be trusted


def read_month_tab(ws, year=None, month=None):
    """Return (reaches, warnings) for rows with any data.

    Each reach carries the completion date recorded per cycle. A date that falls
    outside the tab it sits on means the work is being credited to the wrong
    month, which is precisely what these columns exist to catch, so it is
    reported rather than quietly accepted. Reporting (not failing) is deliberate:
    a stray date is a question for the maintenance team, and the published
    numbers still come from the tab, so one bad cell shouldn't block a refresh.
    """
    reaches, warnings = [], []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not row or len(row) <= REACH_COL:
            continue
        zone = str(row[ZONE_COL]).strip() if row[ZONE_COL] else None
        reach = str(row[REACH_COL]).strip() if row[REACH_COL] else None
        if not zone or not reach:
            continue
        c1, c2 = max_pct(row, C1_COLS), max_pct(row, C2_COLS)
        if c1 is None and c2 is None:
            continue
        dates = {}
        for label, col in (("c1CompletedOn", C1_DATE_COL), ("c2CompletedOn", C2_DATE_COL)):
            val, ok = cell_date(row, col)
            cycle = "Cycle 1" if col == C1_DATE_COL else "Cycle 2"
            if val is None:
                dates[label] = None
            elif not ok:
                dates[label] = None
                warnings.append(f"{ws.title}: {zone} / {reach} {cycle} completion date "
                                f"{val!r} is not a date -- enter it as a date, not text")
            elif year and month and not val.startswith(f"{year:04d}-{month:02d}"):
                dates[label] = val
                warnings.append(f"{ws.title}: {zone} / {reach} {cycle} completed {val}, "
                                f"which is outside {ws.title} {year} -- the work is being "
                                f"credited to the wrong month")
            else:
                dates[label] = val
        reaches.append({"zone": zone, "reach": reach, "cycle1Pct": c1, "cycle2Pct": c2, **dates})
    return reaches, warnings


def main():
    path = resolve_input()
    if not os.path.exists(path):
        print(f"ERROR: File not found: {path}")
        sys.exit(1)

    fn = re.search(r"(\d{4})", os.path.basename(path))
    year = int(fn.group(1)) if fn else None

    print(f"Reading {path}...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Month tabs present, in calendar order; pick the LATEST one that has data.
    month_tabs = sorted(
        [(MONTHS[s.strip().lower()], s) for s in wb.sheetnames if s.strip().lower() in MONTHS]
    )
    active = None
    for num, name in month_tabs:
        reaches, warnings = read_month_tab(wb[name], year, num)
        if reaches:
            active = (num, name, reaches, warnings)

    wb.close()

    if not active or not year:
        print("No month tab with data found (or no year in filename).")
        output = {"reportingMonth": None, "source": os.path.basename(path),
                  "reaches": [], "warnings": []}
    else:
        num, name, reaches, warnings = active
        output = {
            "reportingMonth": {"label": f"{name} {year}", "year": year, "month": num},
            "source": os.path.basename(path),
            # Recorded in the output as well as printed, so a bad date shows up in
            # the committed diff and the refresh digest instead of only in CI logs.
            "warnings": warnings,
            "reaches": reaches,
        }
        print(f"Active month: {name} {year} | {len(reaches)} reaches with data")
        for w in warnings:
            print(f"  WARNING: {w}")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2)
    print(f"Output written to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
