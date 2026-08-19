#!/usr/bin/env python3
"""
Apply the FPA turf input template's structural conventions to a workbook.

Built Aug 2026 after month-end production was going uncredited. The maintenance
workbook ran on Friday week-endings, so a month's last column often closed
before the month did: June's stopped on the 26th, and the Outfall Canals cut
finished June 30 landed in the week ending July 3 on the JULY tab. Because each
tab describes its own month's cycles, that work was credited to neither month.

Two conventions fix it, and this script applies both:

  1. Every month's final weekly column ends on the last calendar day of the
     month, so the following month starts clean on the 1st. Percentages are
     cumulative within a month, so a short final week is valid.
  2. Per-cycle completion-date columns (W = Cycle 1, X = Cycle 2), so the month
     credited can be checked against the date the work actually finished.
     extractTurfData.py warns when the two disagree.

Closed months are left alone by default: rewriting a past month's week-ending
dates would misrepresent what was actually submitted. `--from-month` controls
the cutoff (use 1 when standing up a fresh year).

Usage:
  python3 scripts/buildTurfTemplate.py <input.xlsx> [--out PATH]
                                       [--from-month N] [--year YYYY] [--blank]

  --from-month N  bound months >= N (default: the current month)
  --blank         also clear every entered value, keeping the structure
  --out PATH      output path (default: alongside the input, "-revised" suffix)

Note: the April 29 and June 30 completion dates in the workbook sent to FPA in
August 2026 were transcribed by hand from the weekly notes, not generated here.
This script only applies structure.
"""
import argparse
import calendar
import datetime
import os
import re
import sys
from copy import copy

import openpyxl

MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
}

# 1-indexed here (openpyxl cell coordinates), unlike extractTurfData.py's 0-indexed rows.
LABEL_ROW, DATE_ROW, HEADER_ROW = 4, 5, 6
FIRST_DATA_ROW, LAST_DATA_ROW = 7, 37
WEEK_DATE_COLS = [7, 10, 13, 16, 19]          # G J M P S
LAST_BLOCK = 19                                # S: C1, T: C2, U: notes
C1_COLS = [7, 10, 13, 16, 19]
C2_COLS = [8, 11, 14, 17, 20]
NOTE_COLS = [9, 12, 15, 18, 21]
UPDATED_BY_COL = 22                            # V
C1_DATE_COL, C2_DATE_COL = 23, 24              # W, X

INSTRUCTIONS = [
    "Each month's reporting is bounded by the calendar month. The final column of every "
    "month tab ends on the last day of that month, so work finished on the 30th or 31st "
    "is credited to that month and not to the following one.",
    "Record the actual completion date for each cycle in the C1 and C2 Completion Date "
    "columns. This is what confirms a cycle is credited to the month the work was finished.",
    "Do not insert or delete columns. Add nothing between the weekly blocks; the dashboard "
    "reads fixed column positions and shifted columns are dropped without warning.",
]


def month_tabs(wb):
    return [(MONTHS[s.strip().lower()], s) for s in wb.sheetnames if s.strip().lower() in MONTHS]


def add_completion_columns(ws):
    """Append the two per-cycle date columns, styled like the header beside them."""
    src = ws.cell(row=HEADER_ROW, column=UPDATED_BY_COL)
    for col, label in ((C1_DATE_COL, "C1 Completion Date"), (C2_DATE_COL, "C2 Completion Date")):
        cell = ws.cell(row=HEADER_ROW, column=col)
        cell.value = label
        cell._style = copy(src._style)
        ws.column_dimensions[cell.column_letter].width = 20


def bound_to_month_end(ws, year, month):
    """Point the tab's final weekly column at the last calendar day of the month."""
    last_day = calendar.monthrange(year, month)[1]
    ws.cell(row=LABEL_ROW, column=LAST_BLOCK).value = "Month End:"
    cell = ws.cell(row=DATE_ROW, column=LAST_BLOCK)
    cell.value = datetime.datetime(year, month, last_day)
    ref = ws.cell(row=DATE_ROW, column=WEEK_DATE_COLS[-2])
    cell._style = copy(ref._style)
    if not cell.number_format or cell.number_format == "General":
        cell.number_format = ref.number_format
    for col, label in ((LAST_BLOCK, "Month End C1 %"),
                       (LAST_BLOCK + 1, "Month End C2 %"),
                       (LAST_BLOCK + 2, "Month End Notes")):
        ws.cell(row=HEADER_ROW, column=col).value = label


def update_instructions(wb):
    """Append the conventions to the Instructions tab, skipping any already present."""
    if "Instructions" not in wb.sheetnames:
        return
    ins = wb["Instructions"]
    existing = {str(ins.cell(row=r, column=2).value or "").strip() for r in range(1, 60)}
    row, n = 1, 0
    for r in range(1, 60):                     # first free row after the numbered list
        if ins.cell(row=r, column=1).value is not None:
            row = r + 1
            if isinstance(ins.cell(row=r, column=1).value, (int, float)):
                n = int(ins.cell(row=r, column=1).value)
            elif str(ins.cell(row=r, column=1).value).isdigit():
                n = int(ins.cell(row=r, column=1).value)
    for text in INSTRUCTIONS:
        if text in existing:
            continue
        n += 1
        ins.cell(row=row, column=1).value = n
        ins.cell(row=row, column=2).value = text
        row += 1


def clear_entries(ws):
    for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        if not ws.cell(row=r, column=5).value:
            continue
        for c in C1_COLS + C2_COLS + NOTE_COLS + [UPDATED_BY_COL, C1_DATE_COL, C2_DATE_COL]:
            ws.cell(row=r, column=c).value = None


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("input")
    ap.add_argument("--out")
    ap.add_argument("--year", type=int)
    ap.add_argument("--from-month", type=int, default=datetime.date.today().month,
                    help="bound months >= N (default: current month)")
    ap.add_argument("--blank", action="store_true")
    args = ap.parse_args()

    if not os.path.exists(args.input):
        sys.exit(f"ERROR: file not found: {args.input}")

    year = args.year
    if not year:
        m = re.search(r"(\d{4})", os.path.basename(args.input))
        if not m:
            sys.exit("ERROR: no year in filename; pass --year")
        year = int(m.group(1))

    out = args.out
    if not out:
        base, ext = os.path.splitext(args.input)
        out = f"{base}-revised{ext}"

    wb = openpyxl.load_workbook(args.input)
    bounded = []
    for num, name in sorted(month_tabs(wb)):
        ws = wb[name]
        add_completion_columns(ws)
        if num >= args.from_month:
            bound_to_month_end(ws, year, num)
            bounded.append(name)
        if args.blank:
            clear_entries(ws)
    update_instructions(wb)
    wb.save(out)

    print(f"Wrote {out}")
    print(f"  completion-date columns added to all {len(month_tabs(wb))} month tabs")
    print(f"  month-end bounding applied to: {', '.join(bounded) or '(none)'}")
    if args.blank:
        print("  entered values cleared")


if __name__ == "__main__":
    main()
