#!/usr/bin/env python3
"""
Extract IDIQ contract data from Excel and output JSON for the dashboard.

Reads from: ~/Development/fpa/data/sources/idiq/Contract Master List Revised Beg July 1 2021.xlsx
Outputs to:  ~/Development/fpa/public/data/idiq-contracts.json

Per Director (May 2026), this workbook is refreshed monthly. Drop the new
copy into data/sources/idiq/ (keeping the same filename) and re-run this
script to regenerate the dashboard JSON.
"""

import json
import os
import sys
from collections import OrderedDict
from datetime import datetime

import openpyxl

BASE_DIR = os.path.expanduser("~/Development/fpa/data/sources/idiq")
OUTPUT_PATH = os.path.expanduser("~/Development/fpa/public/data/idiq-contracts.json")

SHEETS = [
    ("2022 IDIQ ", "2022", "2022 IDIQ Contracts"),
    ("2025 IDIQ", "2025", "2025 IDIQ Contracts"),
]


def fmt_date(val):
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str) and val.strip():
        return val.strip()
    return None


def safe_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def extract_pool(ws, pool_id, pool_name):
    contracts_by_service = OrderedDict()
    current = None  # current contract-level record

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue

        contract_num = str(row[0]).strip() if row[0] else None

        # New contract row (has a contract number)
        if contract_num:
            service = str(row[1]).strip() if row[1] else "Uncategorized"
            current = {
                "number": contract_num,
                "consultant": str(row[2]).strip() if row[2] else "",
                "contractDate": fmt_date(row[3]),
                "endDate": fmt_date(row[4]),
                "maximum": safe_float(row[6]),
                "remaining": safe_float(row[7]),
                "service": service,
                "taskOrders": [],
            }
            current["utilized"] = max(0, current["maximum"] - current["remaining"])
            current["utilizationPct"] = (
                round(current["utilized"] / current["maximum"] * 100)
                if current["maximum"] > 0
                else 0
            )

            if service not in contracts_by_service:
                contracts_by_service[service] = []
            contracts_by_service[service].append(current)

        # Task order row (or continuation)
        to_num = str(row[8]).strip() if row[8] else None
        to_status = str(row[9]).strip() if row[9] else None

        if current and to_num and to_status:
            current["taskOrders"].append({
                "number": to_num,
                "status": to_status,
                "description": str(row[10]).strip() if row[10] else "",
                "projectNumber": str(row[11]).strip() if row[11] else "",
                "leveeDistrict": str(row[12]).strip() if row[12] else "",
                "maximum": safe_float(row[13]),
                "costToDate": safe_float(row[14]),
                "startDate": fmt_date(row[15]),
                "endDate": fmt_date(row[16]),
            })

    # Build service type summaries
    service_types = []
    for service, contracts in contracts_by_service.items():
        total_max = sum(c["maximum"] for c in contracts)
        total_utilized = sum(c["utilized"] for c in contracts)
        # Remove service key from individual contracts
        clean_contracts = []
        for c in contracts:
            cc = {k: v for k, v in c.items() if k != "service"}
            clean_contracts.append(cc)

        service_types.append({
            "service": service,
            "contractCount": len(contracts),
            "totalMaximum": total_max,
            "totalUtilized": max(0, total_utilized),
            "utilizationPct": round(total_utilized / total_max * 100) if total_max > 0 else 0,
            "contracts": clean_contracts,
        })

    return {
        "id": pool_id,
        "name": pool_name,
        "serviceTypes": service_types,
    }


def build_summary(pools):
    total_contracts = 0
    total_max = 0
    total_utilized = 0
    active_tos = 0
    completed_tos = 0
    all_services = set()
    all_firms = set()

    for pool in pools:
        for st in pool["serviceTypes"]:
            all_services.add(st["service"])
            total_contracts += st["contractCount"]
            total_max += st["totalMaximum"]
            total_utilized += max(0, st["totalUtilized"])
            for c in st["contracts"]:
                all_firms.add(c["consultant"])
                for to in c["taskOrders"]:
                    if to["status"] == "Active":
                        active_tos += 1
                    elif to["status"] == "Complete":
                        completed_tos += 1

    return {
        "totalContracts": total_contracts,
        "totalMaxValue": total_max,
        "totalUtilized": total_utilized,
        "activeTaskOrders": active_tos,
        "completedTaskOrders": completed_tos,
        "serviceTypes": len(all_services),
        "firms": len(all_firms),
    }


DEFAULT_FILENAME = "Contract Master List Revised Beg July 1 2021.xlsx"


def resolve_input():
    """Input workbook path: CLI arg > IDIQ_INPUT env > default legacy file."""
    if len(sys.argv) > 1:
        return os.path.expanduser(sys.argv[1])
    if os.environ.get("IDIQ_INPUT"):
        return os.path.expanduser(os.environ["IDIQ_INPUT"])
    return os.path.join(BASE_DIR, DEFAULT_FILENAME)


def find_sheet(wb, name):
    """Find a sheet tolerant of trailing/leading whitespace (e.g. '2022 IDIQ ')."""
    if name in wb.sheetnames:
        return wb[name]
    target = name.strip().lower()
    for s in wb.sheetnames:
        if s.strip().lower() == target:
            return wb[s]
    return None


def derive_cycle(pool):
    """Derive a cycle id/name from the earliest contract date year in a pool."""
    years = [
        c["contractDate"][:4]
        for st in pool["serviceTypes"]
        for c in st["contracts"]
        if c.get("contractDate")
    ]
    year = min(years) if years else "current"
    return year, f"{year} IDIQ Contracts"


def load_existing_pools(path):
    """Read the contract pools already on disk, or [] if none/unreadable."""
    if not os.path.exists(path):
        return []
    try:
        with open(path) as f:
            return json.load(f).get("contractPools", [])
    except (ValueError, OSError):
        return []


def upsert_pool(existing, pool):
    """Replace a same-id pool in place, else append. Preserves other cycles
    (e.g. a 2025 upload refreshes the 2025 pool without wiping 2022)."""
    out = list(existing)
    for i, p in enumerate(out):
        if p.get("id") == pool["id"]:
            out[i] = pool
            return out
    out.append(pool)
    return out


def main():
    filepath = resolve_input()
    print(f"Reading {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Two supported shapes:
    #   legacy -> multiple "<year> IDIQ" tabs (the Contract Master List workbook)
    #   upload -> a single flat sheet of all current contracts (FPA's monthly
    #             idiq-contracts_YYYY-MM.xlsx). Whatever the upload contains IS
    #             the dashboard data -- no merging or freezing of prior cycles.
    idiq_tabs = [s for s in wb.sheetnames if "IDIQ" in s.upper()]

    pools = []
    if idiq_tabs:
        for sheet_name, pool_id, pool_name in SHEETS:
            ws = find_sheet(wb, sheet_name)
            if ws is None:
                print(f"  (skipping missing tab: {sheet_name!r})")
                continue
            pools.append(extract_pool(ws, pool_id, pool_name))
    else:
        ws = wb[wb.sheetnames[0]]
        pool = extract_pool(ws, "current", "Current IDIQ Contracts")
        pool["id"], pool["name"] = derive_cycle(pool)
        # Upsert into existing data: refresh this cycle, keep prior cycles (2022)
        # so historic contracts never silently disappear.
        pools = upsert_pool(load_existing_pools(OUTPUT_PATH), pool)

    wb.close()

    for pool in pools:
        contract_count = sum(st["contractCount"] for st in pool["serviceTypes"])
        to_count = sum(
            len(c["taskOrders"])
            for st in pool["serviceTypes"]
            for c in st["contracts"]
        )
        print(f"{pool['name']}: {contract_count} contracts, {to_count} task orders")

    summary = build_summary(pools)
    output = {"contractPools": pools, "summary": summary}

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2)

    print(f"\nSummary: {summary['totalContracts']} contracts, "
          f"{summary['firms']} firms, "
          f"{summary['activeTaskOrders']} active / {summary['completedTaskOrders']} completed TOs, "
          f"${summary['totalMaxValue']:,.0f} total value")
    print(f"Output written to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
