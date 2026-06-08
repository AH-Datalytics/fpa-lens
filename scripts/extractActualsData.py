#!/usr/bin/env python3
"""
Extract FY26 YTD actuals data from the Dashboard Reports workbook.

Reads from: ~/Development/fpa/data/sources/budget/Dashboard_Reports through 03.31.26.xlsm
Outputs to:  ~/Development/fpa/public/data/actuals-fy26.json

Separates Projects from O&M expenses per Jeff Williams' directive:
"separate projects from regular O&M like we did in the FY27 budget because,
at present, the lack of execution for projects really distorts the story"
"""

import calendar
import json
import os
import re
import sys

import openpyxl

# Paths relative to the repo root (scripts/ -> repo) for local + CI parity.
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE_DIR = os.path.join(REPO_ROOT, "data/sources/budget")
OUTPUT_PATH = os.path.join(REPO_ROOT, "public/data/actuals-fy26.json")

FILENAME = "Dashboard_Reports through 03.31.26.xlsm"


def resolve_input():
    """Input workbook path: CLI arg > ACTUALS_INPUT env > legacy default."""
    if len(sys.argv) > 1:
        return os.path.expanduser(sys.argv[1])
    if os.environ.get("ACTUALS_INPUT"):
        return os.path.expanduser(os.environ["ACTUALS_INPUT"])
    return os.path.join(BASE_DIR, FILENAME)


def derive_period(path):
    """Derive the reporting period from a budget-actuals_YYYY-MM filename.
    Falls back to the legacy March 2026 labels if the name has no YYYY-MM."""
    m = re.search(r"(\d{4})-(\d{2})", os.path.basename(path))
    if not m:
        return {
            "period": "FY26 YTD through March 31, 2026",
            "periodLabel": "Jul 2025 - Mar 2026",
            "lastUpdated": "2026-03-31",
            "fiscalYear": 2026,
        }
    year, month = int(m.group(1)), int(m.group(2))
    last_day = calendar.monthrange(year, month)[1]
    fy = year if month <= 6 else year + 1   # fiscal year runs Jul-Jun
    return {
        "period": f"FY{fy % 100} YTD through {calendar.month_name[month]} {last_day}, {year}",
        "periodLabel": f"Jul {fy - 1} - {calendar.month_abbr[month]} {year}",
        "lastUpdated": f"{year}-{month:02d}-{last_day:02d}",
        "fiscalYear": fy,
    }

# Sheet name mappings: key = our entity ID, value = (category sheet, dept sheet)
ENTITY_SHEETS = {
    "wholeEntity": ("Whole Entity", "Whole Entity_Dept"),
    "OLD": ("Summary_Orleans Levee District", "OrleansLeveeDistrict_Dept"),
    "EJLD": ("Summary_East Jefferson Levee Di", "EastJeffLevDistrict_Dept"),
    "LBBLD": ("Summary_LakeBorgneBasinLevDistr", "LakeBorgneBasinLevDist_Dept"),
    "FPAE": ("Summary_FPA-E", "FPA-E_Dept"),
}

ENTITY_LABELS = {
    "wholeEntity": "Whole Entity (All Districts)",
    "OLD": "Orleans Levee District",
    "EJLD": "East Jefferson Levee District",
    "LBBLD": "Lake Borgne Basin Levee District",
    "FPAE": "Flood Protection Authority East",
}

# Revenue categories (in sheet order)
REVENUE_CATS = [
    "Tax Revenue", "Grant Revenue", "Intergovernmental",
    "Rev from EJ", "Rev from LB", "Rev from OL",
    "Interest Income", "Misc Revenue",
]

# Expense categories: O&M (everything except Projects)
OM_EXPENSE_CATS = [
    "Personnel Services", "Training", "Professional Svcs",
    "Contractuals", "Materials & Supplies", "Equipments",
    "Other Charges", "Cost Sharing",
]

PROJECT_CAT = "Projects"


def safe_num(value):
    """Convert a cell value to a number, defaulting to 0."""
    if value is None:
        return 0
    try:
        return round(float(value), 2)
    except (ValueError, TypeError):
        return 0


def safe_variance(value):
    """Convert variance cell to a percentage number or special flag."""
    if value is None:
        return 0
    if isinstance(value, str):
        if "Unbudgeted" in value:
            return "unbudgeted"
        return 0
    try:
        return round(float(value) * 100, 1)  # Convert decimal to percentage
    except (ValueError, TypeError):
        return 0


def extract_category_sheet(ws):
    """Extract revenue, expenses (split O&M vs Projects), sources/uses from a category sheet."""
    # Read all rows: col B=label, C=ytdActual, D=ytdBudget, E=variance, F=totalBudget
    rows = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        label = row[1]
        if label and isinstance(label, str) and label.strip():
            label = label.strip()
            rows[label] = {
                "ytdActual": safe_num(row[2]),
                "ytdBudget": safe_num(row[3]),
                "variancePct": safe_variance(row[4]),
                "totalBudget": safe_num(row[5]),
            }

    # Build revenue
    revenue_cats = []
    for cat in REVENUE_CATS:
        if cat in rows and (rows[cat]["ytdActual"] != 0 or rows[cat]["ytdBudget"] != 0 or rows[cat]["totalBudget"] != 0):
            revenue_cats.append({"name": cat, **rows[cat]})

    revenue_total = rows.get("Total Revenues", {
        "ytdActual": 0, "ytdBudget": 0, "variancePct": 0, "totalBudget": 0,
    })

    # Build O&M expenses
    om_cats = []
    om_actual = 0
    om_budget = 0
    om_total_budget = 0
    for cat in OM_EXPENSE_CATS:
        if cat in rows:
            om_cats.append({"name": cat, **rows[cat]})
            om_actual += rows[cat]["ytdActual"]
            om_budget += rows[cat]["ytdBudget"]
            om_total_budget += rows[cat]["totalBudget"]

    # Compute O&M variance
    om_variance = 0
    if om_budget != 0:
        om_variance = round((om_actual - om_budget) / abs(om_budget) * 100, 1)

    om_total = {
        "ytdActual": round(om_actual, 2),
        "ytdBudget": round(om_budget, 2),
        "variancePct": om_variance,
        "totalBudget": round(om_total_budget, 2),
    }

    # Build projects
    projects_data = rows.get(PROJECT_CAT, {
        "ytdActual": 0, "ytdBudget": 0, "variancePct": 0, "totalBudget": 0,
    })
    projects = {
        "categories": [{"name": "Projects", **projects_data}] if projects_data["totalBudget"] != 0 or projects_data["ytdActual"] != 0 else [],
        "total": projects_data,
    }

    # Grand total expenses
    expense_total = rows.get("Total Expenses", {
        "ytdActual": 0, "ytdBudget": 0, "variancePct": 0, "totalBudget": 0,
    })

    # Sources/Uses
    sources_uses = rows.get("Total SourcesUses", {
        "ytdActual": 0, "ytdBudget": 0, "variancePct": 0, "totalBudget": 0,
    })

    # Net change
    net_change = rows.get("Net Changes in Fund Balance", {
        "ytdActual": 0, "ytdBudget": 0, "variancePct": 0, "totalBudget": 0,
    })

    return {
        "revenue": {
            "categories": revenue_cats,
            "total": {"name": "Total Revenue", **revenue_total},
        },
        "expenses": {
            "operations": {
                "categories": om_cats,
                "total": {"name": "Total O&M", **om_total},
            },
            "projects": projects,
            "grandTotal": {"name": "Total Expenses", **expense_total},
        },
        "sourcesUses": {"name": "Sources & Uses", **sources_uses},
        "netChange": {"name": "Net Change in Fund Balance", **net_change},
    }


def extract_dept_sheet(ws):
    """Extract department-level data from a department sheet."""
    rows = []
    section = None  # "revenue", "expenses", "sourcesUses"

    for row in ws.iter_rows(min_row=5, values_only=True):
        label = row[1]
        if label is None or not isinstance(label, str) or not label.strip():
            continue

        label = label.strip()
        data = {
            "ytdActual": safe_num(row[2]),
            "ytdBudget": safe_num(row[3]),
            "variancePct": safe_variance(row[4]),
            "totalBudget": safe_num(row[5]),
        }

        if label == "Total Revenues":
            section = "post_revenue"
            continue
        elif label == "Total Expenses":
            section = "post_expenses"
            continue
        elif label == "Total SourcesUses":
            section = "done"
            continue
        elif label == "Net Changes in Fund Balance":
            continue

        if section is None:
            # Revenue departments (skip - same info as category sheet)
            continue
        elif section == "post_revenue":
            # Expense departments
            if data["ytdActual"] != 0 or data["ytdBudget"] != 0 or data["totalBudget"] != 0:
                rows.append({"name": label, **data})

    return rows


def main():
    filepath = resolve_input()
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        return

    print(f"Reading {filepath}...")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    entities = {}
    for entity_id, (cat_sheet, dept_sheet) in ENTITY_SHEETS.items():
        print(f"  Extracting {entity_id}...")

        cat_data = extract_category_sheet(wb[cat_sheet])
        dept_data = extract_dept_sheet(wb[dept_sheet])

        entities[entity_id] = {
            "label": ENTITY_LABELS[entity_id],
            **cat_data,
            "departments": dept_data,
        }

    wb.close()

    output = {
        **derive_period(filepath),
        "entities": entities,
    }

    # Summary
    we = entities["wholeEntity"]
    print(f"\n--- Whole Entity Summary ---")
    print(f"Revenue:      ${we['revenue']['total']['ytdActual']:>14,.2f} actual  /  ${we['revenue']['total']['ytdBudget']:>14,.2f} budget  ({we['revenue']['total']['variancePct']:+.1f}%)")
    print(f"O&M Expenses: ${we['expenses']['operations']['total']['ytdActual']:>14,.2f} actual  /  ${we['expenses']['operations']['total']['ytdBudget']:>14,.2f} budget  ({we['expenses']['operations']['total']['variancePct']:+.1f}%)")
    print(f"Projects:     ${we['expenses']['projects']['total']['ytdActual']:>14,.2f} actual  /  ${we['expenses']['projects']['total']['ytdBudget']:>14,.2f} budget  ({we['expenses']['projects']['total']['variancePct']:+.1f}%)")
    print(f"Total Exp:    ${we['expenses']['grandTotal']['ytdActual']:>14,.2f} actual  /  ${we['expenses']['grandTotal']['ytdBudget']:>14,.2f} budget  ({we['expenses']['grandTotal']['variancePct']:+.1f}%)")
    print(f"Net Change:   ${we['netChange']['ytdActual']:>14,.2f}")

    print(f"\nO&M Expenses by Category:")
    for cat in we["expenses"]["operations"]["categories"]:
        v = cat["variancePct"]
        vstr = "⚠ Unbudgeted" if v == "unbudgeted" else f"{v:+.1f}%"
        print(f"  {cat['name']:<22s} ${cat['ytdActual']:>12,.2f} actual  /  ${cat['ytdBudget']:>12,.2f} budget  ({vstr})")

    print(f"\nDepartments:")
    for dept in we["departments"]:
        v = dept["variancePct"]
        vstr = "⚠ Unbudgeted" if v == "unbudgeted" else f"{v:+.1f}%"
        print(f"  {dept['name']:<24s} ${dept['ytdActual']:>12,.2f} actual  /  ${dept['ytdBudget']:>12,.2f} budget  ({vstr})")

    # Write output
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2)

    print(f"\nOutput written to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
