#!/usr/bin/env python3
"""
Extract safety event data from Excel files and output anonymized JSON.

Reads from: ~/Development/fpa/data/sources/safety-event-logs/*.xlsx
Outputs to:  ~/Development/fpa/public/data/safety-events.json

Per Safety Officer (Jamal) Apr 2026 update, each row is classified by the
fill color of the Recordable cell:

  RED (FFFF0000)        -> N/A         -> excluded from ALL metrics
  LIGHT BLUE (FF00B0F0) -> No-Fault    -> excluded from performance and
                                          year-over-year metrics, but kept
                                          in total event tracking
  No fill (00000000)    -> At-Fault    -> further classified as:
       Recordable=Yes -> OSHA Recordable accident
       Damage=Yes     -> Property damage event (FPA and/or private)

Anonymization: strips employee names, group codes, locations, unit numbers,
make/model, descriptions. Only retains date, event type, classification,
and boolean flags (lost time, injury, damage, private-property damage).
"""

import json
import os
import sys
from collections import defaultdict
from datetime import datetime

import openpyxl

# Paths relative to the repo root (scripts/ -> repo) for local + CI parity.
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE_DIR = os.path.join(REPO_ROOT, "data/sources/safety-event-logs")
OUTPUT_PATH = os.path.join(REPO_ROOT, "public/data/safety-events.json")

# Frozen, already-anonymized history of the years that no longer change. The
# weekly pipeline reads this as its base and only refreshes the current year
# from the SharePoint upload, so the raw logs (which contain employee PII) never
# need to live in the repo / CI.
HISTORY_PATH = os.path.join(REPO_ROOT, "scripts/safety-history.json")
HISTORY_THROUGH_YEAR = 2025
HISTORICAL_FILES = [
    ("Event Log 2022.xlsx", "Events"),
    ("2023 Event Log.xlsx", "Sheet2"),
    ("2024 Event Log.xlsx", "Sheet1"),
    ("2025 Event Log.xlsx", "Sheet1"),
]

# File configurations: (filename, sheet_name)
# As of the Apr 2026 reclassified workbooks, all years share the same layout:
#   1:Date | 2:Time | 3:Employee | 4:Group | 5:Event Type
#   6:Event Description | 7:Root Cause | 8:Corrective Action | 9:Location
#   10:Injury/Body Part | 11:Recordable | 12:Lost Time | 13:Damage
#   14:FPA Unit No. | 15:Make Model/Description | 16:Damage to Private Property
#   17:Photo | 18:Notes
FILES = [
    ("Event Log 2022.xlsx", "Events"),
    ("2023 Event Log.xlsx", "Sheet2"),
    ("2024 Event Log.xlsx", "Sheet1"),
    ("2025 Event Log.xlsx", "Sheet1"),
    ("2026 Event Log.xlsx", "Sheet1"),
]

COL_DATE = 1
COL_EVENT_TYPE = 5
COL_INJURY = 10
COL_RECORDABLE = 11
COL_LOST_TIME = 12
COL_DAMAGE = 13
COL_PRIVATE_PROP = 16

# Cell fill colors used by the Safety Officer to classify rows.
COLOR_NA = "FFFF0000"          # red    -> exclude entirely
COLOR_NO_FAULT = "FF00B0F0"    # cyan   -> total tracking only
COLOR_AT_FAULT = ("00000000", "", None)  # white / no fill -> counted everywhere


def parse_yes(value):
    """Treat any non-empty string that isn't an explicit "no" as Yes."""
    if value is None:
        return False
    s = str(value).strip().lower()
    if s in ("", "no", "none", "n/a", "na", "unknown", "unk", "tbd", "pending"):
        return False
    return True


def parse_injury(value):
    """Injury column is free text. Empty / 'No' / 'N/A' means no injury."""
    if value is None:
        return False
    s = str(value).strip().lower()
    if s in ("", "no", "none", "n/a", "na", "unknown", "unk"):
        return False
    return True


def parse_date(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def cell_fill_color(cell):
    """Return the foreground fill color of a cell, normalized to upper-case hex."""
    fill = cell.fill
    if fill is None or fill.fgColor is None:
        return None
    val = fill.fgColor.value
    if val is None:
        return None
    return str(val).upper()


def classify_row(recordable_fill):
    """
    Classify a row based on the fill color of its Recordable cell.

    Returns one of:
      "excluded"   -- N/A row, drop entirely
      "no-fault"   -- counted in totals only
      "at-fault"   -- counted everywhere; further sub-classified by columns
    """
    if recordable_fill == COLOR_NA:
        return "excluded"
    if recordable_fill == COLOR_NO_FAULT:
        return "no-fault"
    return "at-fault"


def find_data_sheet(wb):
    """Pick the sheet whose A1 header is 'Date' (the event-log layout)."""
    for s in wb.sheetnames:
        a1 = wb[s].cell(row=1, column=1).value
        if a1 and str(a1).strip().lower() == "date":
            return wb[s]
    return wb[wb.sheetnames[0]]


def events_from_file(filepath, sheet_name=None):
    """Return the anonymized event list from one workbook. Sheet is auto-detected
    when not given, since uploads vary ('Sheet1', 'Events', ...)."""
    if not os.path.exists(filepath):
        print(f"WARNING: File not found: {filepath}")
        return []

    # Need fill-color formatting, so don't use read_only mode.
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else find_data_sheet(wb)

    events = []
    for r in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=r, column=COL_DATE)
        date_val = parse_date(date_cell.value)
        if date_val is None:
            continue

        recordable_cell = ws.cell(row=r, column=COL_RECORDABLE)
        classification = classify_row(cell_fill_color(recordable_cell))
        if classification == "excluded":
            continue

        event_type = "Not categorized"
        v = ws.cell(row=r, column=COL_EVENT_TYPE).value
        if v is not None and str(v).strip():
            event_type = str(v).strip()

        is_recordable = parse_yes(recordable_cell.value)
        is_lost_time = parse_yes(ws.cell(row=r, column=COL_LOST_TIME).value)
        has_damage = parse_yes(ws.cell(row=r, column=COL_DAMAGE).value)
        has_private_property_damage = parse_yes(ws.cell(row=r, column=COL_PRIVATE_PROP).value)
        has_injury = parse_injury(ws.cell(row=r, column=COL_INJURY).value)

        # Sub-classification for at-fault rows: OSHA-recordable accident
        # (Recordable=Yes), a damage-only event, or other. No-fault rows always
        # carry property damage, by Safety Officer convention.
        if classification == "at-fault":
            if is_recordable:
                sub = "osha-recordable"
            elif has_damage or has_private_property_damage:
                sub = "damage"
            else:
                sub = "other"
        else:
            sub = "no-fault"

        events.append({
            "date": date_val.strftime("%Y-%m-%d"),
            "month": date_val.month,
            "year": date_val.year,
            "monthName": date_val.strftime("%B"),
            "eventType": event_type,
            "classification": sub,         # osha-recordable | damage | other | no-fault
            "isAtFault": classification == "at-fault",
            "isRecordable": is_recordable,
            "isLostTime": is_lost_time,
            "hasInjury": has_injury,
            "hasDamage": has_damage,
            "hasPrivatePropertyDamage": has_private_property_damage,
        })

    wb.close()
    print(f"  {os.path.basename(filepath)} [{ws.title}]: {len(events)} events")
    return events


def extract_events(files=None):
    """Aggregate anonymized events across (filename, sheet) pairs in BASE_DIR,
    defaulting to all logs (legacy full rebuild)."""
    files = files if files is not None else FILES
    all_events = []
    for filename, sheet_name in files:
        all_events.extend(events_from_file(os.path.join(BASE_DIR, filename), sheet_name))
    return all_events


# Safety Officer's official annual summary for the closed years. These
# authoritative totals (revised by Jamal Dortch, June 2026) supersede the
# event-log-derived classification for 2022-2025, which differs slightly because
# of how individual rows were tagged. Only the closed years are overridden; the
# live year (2026) stays fully event-log-derived. Fields not listed here
# (no-fault, lost time, property damage) keep their derived values.
OFFICIAL_YEARLY_OVERRIDES = {
    2022: {"accidents": 12, "incidents": 32, "injuries": 20},
    2023: {"accidents": 9, "incidents": 19, "injuries": 12},
    2024: {"accidents": 6, "incidents": 37, "injuries": 11},
    2025: {"accidents": 6, "incidents": 24, "injuries": 11},
}

# Safety Officer's official MONTHLY accident/incident split for the closed years,
# from Jamal Dortch's June 2026 monthly breakdown ("Lens Monthly Safety Numbers"
# workbook). His team re-reviewed the 2022-2024 records and re-cut the monthly
# data so each month reconciles to the revised annual totals above; the
# event-log-derived split differed (notably 2022 and 2024, where row-level
# tagging diverged from the final classification). Each value is
# (accidents, incidents) per month (1=Jan ... 12=Dec); every year sums to its
# OFFICIAL_YEARLY_OVERRIDES accidents/incidents. No-fault counts are not part of
# his breakdown and keep their event-log-derived monthly values; injuries are
# tracked only at the year level, so they are not carried here. The live year
# (2026) is intentionally absent -- it stays fully event-log-derived so weekly
# refreshes keep updating it.
OFFICIAL_MONTHLY_OVERRIDES = {
    2022: {1: (1, 2), 2: (0, 2), 3: (1, 3), 4: (1, 5), 5: (2, 3), 6: (1, 8), 7: (1, 0), 8: (2, 1), 9: (0, 3), 10: (3, 2), 11: (0, 1), 12: (0, 2)},
    2023: {1: (0, 1), 2: (1, 1), 3: (1, 3), 4: (0, 1), 5: (2, 2), 6: (1, 2), 7: (1, 2), 8: (1, 3), 9: (1, 1), 10: (1, 0), 11: (0, 3), 12: (0, 0)},
    2024: {1: (0, 3), 2: (0, 1), 3: (0, 6), 4: (0, 1), 5: (0, 4), 6: (0, 3), 7: (2, 6), 8: (0, 3), 9: (1, 1), 10: (0, 5), 11: (2, 1), 12: (1, 3)},
    2025: {1: (1, 0), 2: (0, 2), 3: (0, 3), 4: (0, 0), 5: (0, 2), 6: (0, 1), 7: (0, 3), 8: (3, 2), 9: (0, 5), 10: (2, 4), 11: (0, 1), 12: (0, 1)},
}


def build_output(events):
    # ---- yearlyTotals (excluding N/A by construction; further breakdowns) ----
    # Reporting framing: each event is an Accident, Incident, or No-Fault Event.
    #   Accident   = at-fault, OSHA-recordable injury
    #   Incident   = at-fault, no recordable injury (property damage or near-miss)
    #   No-Fault   = employee not at fault (Safety Officer's cyan classification)
    yearly = defaultdict(lambda: {
        "totalEvents": 0,
        "accidents": 0,              # at-fault OSHA recordable
        "incidents": 0,              # at-fault, not recordable (damage + minor)
        "noFault": 0,
        "lostTime": 0,
        "injuries": 0,
        "propertyDamageFpa": 0,
        "propertyDamagePrivate": 0,
    })
    for e in events:
        y = yearly[e["year"]]
        y["totalEvents"] += 1
        if e["classification"] == "osha-recordable":
            y["accidents"] += 1
        elif e["classification"] in ("damage", "other"):
            y["incidents"] += 1
        elif e["classification"] == "no-fault":
            y["noFault"] += 1
        if e["isLostTime"]:
            y["lostTime"] += 1
        if e["hasInjury"]:
            y["injuries"] += 1
        if e["hasDamage"]:
            y["propertyDamageFpa"] += 1
        if e["hasPrivatePropertyDamage"]:
            y["propertyDamagePrivate"] += 1

    yearly_totals = []
    for year in sorted(yearly.keys()):
        entry = {"year": year}
        entry.update(yearly[year])
        # Overlay the Safety Officer's official totals for closed years. The
        # public page derives "Total Events" as accidents + incidents, so
        # overriding those two (plus injuries) is enough to show his numbers;
        # we also keep the stored totalEvents field consistent.
        override = OFFICIAL_YEARLY_OVERRIDES.get(year)
        if override:
            entry.update(override)
            entry["totalEvents"] = entry["accidents"] + entry["incidents"] + entry["noFault"]
        yearly_totals.append(entry)

    # ---- monthlyData ----
    monthly = defaultdict(lambda: {
        "accidents": 0,
        "incidents": 0,
        "noFault": 0,
    })
    for e in events:
        key = (e["year"], e["month"])
        if e["classification"] == "osha-recordable":
            monthly[key]["accidents"] += 1
        elif e["classification"] in ("damage", "other"):
            monthly[key]["incidents"] += 1
        elif e["classification"] == "no-fault":
            monthly[key]["noFault"] += 1

    # Overlay the Safety Officer's official monthly accident/incident counts for
    # the closed years so the per-month chart reconciles with the headline annual
    # totals. noFault stays event-log-derived (not part of his breakdown), and a
    # month with no logged events still picks up its official accident/incident
    # split via the defaultdict (noFault defaults to 0).
    for year, months in OFFICIAL_MONTHLY_OVERRIDES.items():
        for month, (accidents, incidents) in months.items():
            m = monthly[(year, month)]
            m["accidents"] = accidents
            m["incidents"] = incidents

    monthly_data = []
    for (year, month) in sorted(monthly.keys()):
        m = monthly[(year, month)]
        monthly_data.append({
            "year": year,
            "month": month,
            "accidents": m["accidents"],
            "incidents": m["incidents"],
            "noFault": m["noFault"],
        })

    # ---- eventTypes ----
    type_counts = defaultdict(lambda: {
        "count": 0,
        "accidents": 0,
        "incidents": 0,
        "noFault": 0,
    })
    for e in events:
        if e["eventType"] == "Not categorized":
            continue
        t = type_counts[e["eventType"]]
        t["count"] += 1
        if e["classification"] == "osha-recordable":
            t["accidents"] += 1
        elif e["classification"] in ("damage", "other"):
            t["incidents"] += 1
        elif e["classification"] == "no-fault":
            t["noFault"] += 1

    event_types = []
    for etype in sorted(type_counts.keys()):
        entry = {"type": etype}
        entry.update(type_counts[etype])
        event_types.append(entry)

    # ---- recentEvents (anonymized) ----
    recent_events = sorted(events, key=lambda e: e["date"], reverse=True)
    recent_events = [
        {
            "date": e["date"],
            "month": e["monthName"],
            "year": e["year"],
            "eventType": e["eventType"],
            "classification": e["classification"],
            "isAtFault": e["isAtFault"],
            "isLostTime": e["isLostTime"],
            "hasInjury": e["hasInjury"],
            "hasDamage": e["hasDamage"],
            "hasPrivatePropertyDamage": e["hasPrivatePropertyDamage"],
        }
        for e in recent_events
    ]

    return {
        "yearlyTotals": yearly_totals,
        "monthlyData": monthly_data,
        "eventTypes": event_types,
        "recentEvents": recent_events,
    }


def build_history():
    """Freeze the static historical years to an anonymized JSON (no PII). Run
    this once a year after the prior year's log is final."""
    events = extract_events(HISTORICAL_FILES)
    with open(HISTORY_PATH, "w") as f:
        json.dump({"throughYear": HISTORY_THROUGH_YEAR, "events": events}, f, indent=2)
    print(f"Wrote {len(events)} historical events (through {HISTORY_THROUGH_YEAR}) to {HISTORY_PATH}")


def load_history():
    if not os.path.exists(HISTORY_PATH):
        return HISTORY_THROUGH_YEAR, []
    with open(HISTORY_PATH) as f:
        h = json.load(f)
    return h.get("throughYear", HISTORY_THROUGH_YEAR), h.get("events", [])


def write_output(events):
    output = build_output(events)
    print("\nYearly totals:")
    for yt in output["yearlyTotals"]:
        print(f"  {yt['year']}: total={yt['totalEvents']} "
              f"accidents={yt['accidents']} incidents={yt['incidents']} "
              f"no-fault={yt['noFault']} lost-time={yt['lostTime']} "
              f"injuries={yt['injuries']} FPA damage={yt['propertyDamageFpa']} "
              f"private damage={yt['propertyDamagePrivate']}")
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2)
    print(f"\nOutput written to {OUTPUT_PATH}")


def main():
    args = sys.argv[1:]

    # Mode 1: rebuild the frozen anonymized history (run after a year closes).
    if "--build-history" in args:
        build_history()
        return

    # Mode 2: pipeline -- a current-year upload path was passed. Combine the
    # frozen history with the upload's newer-year events (no double counting).
    upload = next((a for a in args if not a.startswith("-")), None)
    if upload:
        through_year, hist_events = load_history()
        print(f"Loaded {len(hist_events)} historical events (through {through_year})")
        upload_events = events_from_file(upload)
        current = [e for e in upload_events if e["year"] > through_year]
        print(f"Using {len(current)} of {len(upload_events)} upload events (year > {through_year})")
        write_output(hist_events + current)
        return

    # Mode 3: legacy full rebuild from all local logs in BASE_DIR.
    print("Extracting safety event data (full rebuild)...")
    write_output(extract_events())


if __name__ == "__main__":
    main()
