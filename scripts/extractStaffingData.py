#!/usr/bin/env python3
"""
Extract current staffing headcount from the monthly workbook into JSON.

Reads:  the FPA Lens Staffing Headcount workbook (Unit / Department / Filled /
        Vacant / Total layout).
Writes: public/data/staffing.json

Only the *current* filled/vacant counts are produced here. Budgeted capacity
(`full`) and the Green/Amber/Red thresholds are Director-set policy and stay in
the dashboard's siteData.ts; the page overlays these current counts on top.

Department labels are emitted exactly as they appear so they line up with the
dashboard's department labels (Maintenance, Operations, Engineering, Police,
Executive, Finance, Human Resources, Information Technology).

Usage:
  python3 scripts/extractStaffingData.py [path-to-staffing.xlsx]
  STAFFING_INPUT=<path> python3 scripts/extractStaffingData.py
"""
import calendar
import json
import os
import re
import sys

import openpyxl

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE_DIR = os.path.join(REPO_ROOT, "data/sources/staffing")
OUTPUT_PATH = os.path.join(REPO_ROOT, "public/data/staffing.json")
DEFAULT_FILENAME = "FPA Lens Staffing Headcount.xlsx"

SKIP_LABELS = {"department", "total", "unit", ""}


def resolve_input():
    if len(sys.argv) > 1:
        return os.path.expanduser(sys.argv[1])
    if os.environ.get("STAFFING_INPUT"):
        return os.path.expanduser(os.environ["STAFFING_INPUT"])
    return os.path.join(BASE_DIR, DEFAULT_FILENAME)


def as_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def derive_as_of(path):
    m = re.search(r"(\d{4})-(\d{2})", os.path.basename(path))
    if not m:
        return None
    return f"{calendar.month_name[int(m.group(2))]} {m.group(1)}"


def main():
    path = resolve_input()
    if not os.path.exists(path):
        print(f"ERROR: File not found: {path}")
        sys.exit(1)

    print(f"Reading {path}...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    departments = {}
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 4:
            continue
        label = str(row[1]).strip() if row[1] else ""
        filled = as_int(row[2])
        if label.lower() in SKIP_LABELS or filled is None:
            continue
        vacant = as_int(row[3]) or 0
        total = as_int(row[4]) if len(row) > 4 and as_int(row[4]) is not None else filled + vacant
        departments[label] = {"filled": filled, "vacant": vacant, "total": total}

    wb.close()

    total_filled = sum(d["filled"] for d in departments.values())
    total_vacant = sum(d["vacant"] for d in departments.values())
    output = {
        "asOf": derive_as_of(path),
        "source": os.path.basename(path),
        "headcount": {
            "total": total_filled + total_vacant,
            "filled": total_filled,
            "vacancies": total_vacant,
        },
        "departments": departments,
    }

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2)

    print(f"Departments: {', '.join(departments)}")
    print(f"Total positions {output['headcount']['total']} | "
          f"filled {total_filled} | vacancies {total_vacant}")
    print(f"Output written to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
